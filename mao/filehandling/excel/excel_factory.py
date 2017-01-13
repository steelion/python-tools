__author__ = 'Ofner Mario'

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, fills

from logging import getLogger

import pprint
import os.path


logger = getLogger(__name__)

class ExcelFactory():

    workbook = None
    filename = ""
    ws = None


    def create_new_workbook(self, _filename = None):
        logger.debug("Create new Workbook")
        self.workbook = Workbook()

        if _filename is not None:
            self.filename = _filename

        self.save_workbook()
        logger.debug("Finished creating new Workbook")


    def save_workbook(self):
        if self.workbook is not None:
            if self.filename != "":
                self.workbook.save(filename = self.filename)
                logger.debug("Workbook saved to: " + str(self.filename))
            else:
                logger.warning("No Filename set - Workbook not saved")
        else:
            logger.warning("No Workbook created - Nothing to save")


    def load_workbook(self, _filename = None):
        if _filename is not None:
            logger.debug("Load Workbook using filename: " + str(_filename))

            if not os.path.isfile(_filename):
                self.create_new_workbook(_filename = _filename)
            else:
                logger.debug("File exists, loading Wockbook")
                self.workbook = load_workbook(filename = _filename)

            self.filename = _filename
            logger.debug("Loading Workbook done")


    def create_worksheet(self, _name):
        self.workbook.create_sheet(title = _name)
        self.change_worksheet(_name)


    def change_worksheet(self, _name):
        if _name in self.workbook.get_sheet_names():
            self.ws = self.workbook.get_sheet_by_name(_name)
            return True
        return False

    def get_worksheet_name(self):
        return self.ws


    def remove_worksheet(self, _name):
        if _name in self.workbook.get_sheet_names():
            self.workbook.remove_sheet(self.workbook.get_sheet_by_name(_name))
            logger.debug("worksheet '" + _name + "' removed")
            return True
        return False

    def merge_cells(self, start_row, start_column, end_row, end_column):
        self.ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)


    def get_cell_value(self, row, col):
        col_letter = get_column_letter(col)
        return self.ws[str(col_letter) + str(row)].value


    def set_cell_value(self, row, col, value):
        col_letter = get_column_letter(col)
        self.ws[str(col_letter) + str(row)] = value


    def set_row_values(self, row, row_data_values):
        col=1
        for i in row_data_values:
            if i is not None:
                self.ws[get_column_letter(col)+str(row)]=i
            col+=1


    def set_row_format(self, row, row_format):
        col=1
        for i in row_format:
            if i is not None:
                self.ws[get_column_letter(col)+str(row)].number_format=i
            col+=1

    def set_additional_styles(self, row, additional_format_list):
        col=1
        for additional_format in additional_format_list:
            # logger.debug("column: "+str(col)+", format: "+str(additional_format))

            if additional_format is not None:
                alignment = None
                if additional_format == "text_wrap":
                    alignment = Alignment(wrap_text=True)
                elif additional_format == "text_wrap; rotated":
                    alignment = Alignment(wrap_text=True, text_rotation=90, horizontal="center")

                if alignment is not None:
                    self.set_cell_alignment(row, col, alignment)
            col+=1

    def set_cell_text_style(self, row, col, text_style):

        alignment = None

        if text_style == "h_left":
            alignment = Alignment(horizontal="left")
        elif text_style == "h_center":
            alignment = Alignment(horizontal="center")
        elif text_style == "h_right":
            alignment = Alignment(horizontal="right")

        elif text_style == "v_top":
            alignment = Alignment(vertical="top")
        elif text_style == "v_center":
            alignment = Alignment(vertical="center")
        elif text_style == "v_bottom":
            alignment = Alignment(vertical="bottom")

        elif text_style == "v_center;h_center":
            alignment = Alignment(vertical="center", horizontal="center")
        elif text_style == "v_center;h_right":
            alignment = Alignment(vertical="center", horizontal="right")


        else:
            logger.error("Unknown text_style: "+repr(text_style))

        if alignment is not None:
            self.set_cell_alignment(row, col, alignment)


    def set_cell_font_style(self, row, col, font_style):
        if font_style is not None:
            font = Font(
                name=font_style["name"],
                size=font_style["size"],
                bold=font_style["bold"],
                italic=font_style["italic"],
                vertAlign=font_style["vertAlign"],
                underline=font_style["underline"],
                strike=font_style["strike"],
                color=font_style["color"]
                )
            cell = self.ws.cell(column=col, row=row)
            cell.font = font

    def set_cell_alignment(self, row, col, cell_alignment):
        cell = self.ws.cell(column=col, row=row)
        cell.alignment = cell_alignment

    def set_cell_bgcolor(self, row, col, bgcolor_hex, patterntype = None):
        if patterntype is None:
            patterntype = fills.FILL_SOLID

        fill = PatternFill(patternType=patterntype, start_color=bgcolor_hex)
        self.set_cell_fill(row, col, fill)


    def set_cell_fill(self, row, col, pattern_fill):
            cell = self.ws.cell(column=col, row=row)
            cell.fill = pattern_fill


    def set_column_width(self, dimension_width_list):
        col=1
        for col_width in dimension_width_list:
            if col_width is not None:
                self.ws.column_dimensions[get_column_letter(col)].width = col_width
            col+=1


    def set_font(self, row, font):
        for col in range(1, self.ws.max_column+1):
            cell = self.ws.cell(column=col, row=row)
            cell.font = font


    def set_row_heigth(self, row, dimension_row_height):
        self.ws.row_dimensions[row].height = dimension_row_height


    def apply_styles_to_row(self, row = None, number_format_list = None, additional_format_list = None, dimension_row_height = None, font_description = None):
        if row is None:
            rownum=self.get_last_row()
        else:
            rownum = row

        if number_format_list is not None:
            self.set_row_format(rownum, number_format_list)

        if dimension_row_height is not None:
            self.set_row_heigth(rownum, dimension_row_height)

        if additional_format_list is not None:
            self.set_additional_styles(rownum, additional_format_list)

        if font_description is not None:
            font = Font(
                name=font_description["name"],
                size=font_description["size"],
                bold=font_description["bold"],
                italic=font_description["italic"],
                vertAlign=font_description["vertAlign"],
                underline=font_description["underline"],
                strike=font_description["strike"],
                color=font_description["color"]
                )
            self.set_font(rownum, font)


    def append_row(self, row_data_values):
        rownum=self.get_last_row()+1
        self.set_row_values(rownum, row_data_values)


    def get_last_row(self):
        return self.ws.max_row


    def get_last_column(self):
        return self.ws.max_column


    def get_all_values_for_column(self, col):
        table_rowlength = self.get_last_row()
        column_cell_values = []

        for col in self.ws.iter_cols(min_col=col, min_row=2, max_col=col, max_row=table_rowlength):
            for cell in col:
                column_cell_values.append(cell.value)

        return column_cell_values


    def get_all_values_for_row(self, row):
        table_rowwidth = self.get_last_column()
        row_cell_values = []

        for col in self.ws.iter_cols(min_col=1, min_row=row, max_col=table_rowwidth, max_row=row):
            for cell in col:
                row_cell_values.append(cell.value)

        return row_cell_values


    def get_table_header(self):
        return self.get_all_values_for_row(1)


    def draw_table_border( self, min_col, min_row, max_col, max_row, border_only = False):

        bcb = '00000000' # border color black
        bwh = 'hair'
        bwt = 'thin'

        bs_thin = Side(border_style=bwt, color=bcb)
        bs_hair = Side(border_style=bwh, color=bcb)
        bs_none = Side(border_style=None, color=bcb)

        bs_grid = bs_hair
        if border_only:
            bs_grid = bs_none

        for col in range(min_col, max_col+1):
            for row in range(min_row, max_row+1):

                bsl = bs_thin if col == min_col else bs_grid
                bst = bs_thin if row == min_row else bs_grid
                bsr = bs_thin if col == max_col else bs_grid
                bsb = bs_thin if row == max_row else bs_grid

                cell_border = Border(left=bsl, right=bsr, top=bst, bottom=bsb)

                cell = self.ws.cell(column=col, row=row)
                cell.border = cell_border


    def set_number_format( _worksheet, _table_begin_cell, _table_end_cell, _number_formats, _has_header = False ):

        begin_column = _worksheet[_table_begin_cell].column
        begin_column_index = column_index_from_string(begin_column)
        begin_row = _worksheet[_table_begin_cell].row
        end_column = _worksheet[_table_end_cell].column
        end_column_index = column_index_from_string(end_column)
        end_row = _worksheet[_table_end_cell].row

        # print("begin: " + begin_column + ", index: " + str(begin_column_index) + ", row: " + str(begin_row))
        # print("end: " + end_column + ", index: " + str(end_column_index) + ", row: " + str(end_row))

        styling_header = _has_header
        styling_first_table_row = True
        styling_index = 0
        styling_max_column = end_column_index - begin_column_index

        for table_row in range(begin_row, end_row+1):
            if styling_header:
                styling_header = False
                continue

            for table_col in range(begin_column_index, end_column_index+1):
                cell = _worksheet.cell(column=table_col, row=table_row)
                cell.number_format = _number_formats[get_column_letter(table_col)]
                # print('"'+ get_column_letter(table_col) + '"'+str(table_row)+': "' + str(cell.number_format)+'",')

    def colorize_column( _worksheet, _col_color, _col_reference, _color_level = 0.8 ):
        """
        _col_color: The Column Letter of the Column that should be colorized
        _col_reference: The Column Letter of the Column that is used as reference
        _color_level: (float) This "Level" is used for the orange color (Default: 0.8)"""

        pFill_orange = PatternFill(patternType=fills.FILL_SOLID, start_color='FCD5B4')
        pFill_red = PatternFill(patternType=fills.FILL_SOLID, start_color='E6B9B8')

        if _color_level is None:
            level = 0.8

        try:
            level = float(_color_level)
        except (ValueError, TypeError) as e:
            print("Unable to set color_level -> defaults to 0.8")
            level = float(0.8)

        col1_index = column_index_from_string(_col_color)
        col_ref_index = column_index_from_string(_col_reference)

        table_rows = _worksheet.max_row

        for row_num in range(2, table_rows+1):
            col1_value = _worksheet[str(_col_color)+str(row_num)].value
            col_ref_value = _worksheet[str(_col_reference)+str(row_num)].value

            try:
                c1 = float(col1_value)
                c2 = float(col_ref_value)
            except (ValueError, TypeError) as e:
                continue

            if c1 > (c2*level):
                if c1 > c2:
                    _worksheet[str(_col_color)+str(row_num)].fill = pFill_red
                else:
                    _worksheet[str(_col_color)+str(row_num)].fill = pFill_orange


    def set_default_font( _worksheet ):
        default_font = Font(name='Calibri', size=9, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
        table_rows = _worksheet.max_row
        table_width = _worksheet.max_column
        for table_row in range(1, table_rows+1):
            for table_col in range(1, table_width+1):
                cell = _worksheet.cell(column=table_col, row=table_row)
                cell.font = default_fon


